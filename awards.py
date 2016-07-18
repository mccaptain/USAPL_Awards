from sys import argv
import os

import openpyxl
from collections import deque
from collections import namedtuple
lifter = namedtuple("lifter", "name divw place wilks event")

script, file = argv
wb = openpyxl.load_workbook(file,data_only=True)
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Lifting')
maxcol = sheet.max_column
maxrow = sheet.max_row

#clear warnings
os.system('cls')

print  """
                                                                        
          ''';''+;';+                                                   
         ;;;+;;'';;++                                                   
        `;;+;;;+;;'++                                                   
        ;;;+;;'';;+++                                                   
        ;;+;;;+;;;+++                                                   
       :;;+;;;+;;++.::::,      :::::   ,:::::::::::::        .::::      
       ;;;+;;+;;;++`;;;;      `;;;;` `;;;;;;;;;;;;;;,       ,;;;;;.     
      `;;+;;;+;;'++;;;;;      :;;;;  ;;;;;;;;;;;;;;;       ,;;;;;;;     
      :;;+;;;';;++`;;;;,      ;;;;: ;;;;;                 :;;;;;;;;     
      ;;;+;;+;;;++`;;;;      `;;;;  ;;;;;;;;;;;;;,       :;;;;:;;;;:    
     `;;+;;;+;;'+';;;;;      :;;;;  ;;;;;;;;;;;;;;`     :;;;;. ;;;;;    
     :;;+;;;+;;++`;;;;,      ;;;;:  `;;;;;;;;;;;;;:    :;;;;.  ;;;;;`   
     ;;;+;;+;;;++.;;;;      `;;;;            `;;;;,   ;;;;;`   .;;;;;   
     ;;'';;+;;;++:;;;;;;;;;;;;;;; ,;;;;;;;;;;;;;;;   ;;;;;;;;;;:;;;;;   
    .;;+;;;+;;+++`;;;;;;;;;;;;;;  ;;;;;;;;;;;;;;;   ;;;;;;;;;;: :;;;;,  
  ..:;;+;;;';;+++,:;;;;;;;;;;;,   ;;;;;;;;;;;;;:   ;;;;;;;;;;:  `;;;;;  
  ;;;;;+;;+;;;++,::::::::::::::::::::::::::::::::::::::::::::::::;;;;;` 
 .;;;;+';;+;;;+.;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;: 
 .++;;+;;;+;;++,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,, 
   :;;+;;;+;;+++ ++++ ++++`+ .+ ',++++ ++++ +   + ++++++++:+ ++  + ++++ 
   ;;;+;;';;;+++;;`,+''``+`+ ++ + +   ;;``+`+   + +   `:'` +`++ ,+''  + 
   ;;'+;;+;;;++'+..+:+   + +.:+;,;+++ +,.';':  ,':+++  +. :';;+;+.+ +++ 
   +++'+++++;++`++++ +  ,+ ++ ++ +    ++++ +   +.+,,,  +  +`+ ;++ + `': 
  `+++;+'+++'++,+   `++++` +, +, ++++.+  + +++,+ +    `+  + + `++`++++  
  ,+++'+;++++++'                                                        
  ;+'+++''++++:;                                                        
  '+;+++';++++++                                                        
  ++;++++;+'++++                                                        
  ++''+++'+;+++.                                                        
  ++''+++++;+++                                                         
  +++'+++++;+++                                                         
  +++'+++++;++.                                                         
  ++''+++'+;++                                                          
  ;;;+;;':;;+                                                           
                                                                        

"""
namec = 2
divweightc = 80
placec = 82
wilksc = 53
eventc = 31
theLifters = []

for i in range( 13, maxrow ):
    thecell = sheet.cell( row=i, column=namec )
    thename = thecell.value
    if thename != None:
        thelifter = lifter( thename,
                            (sheet.cell( row=i, column=divweightc ).value, sheet.cell( row=i, column=eventc ).value),
                            sheet.cell( row=i, column=placec ).value,
							sheet.cell( row=i, column=wilksc ).value,
							sheet.cell( row=i, column=eventc ).value )
	theLifters.append( thelifter )
awards = dict()
for hum in theLifters:
	if hum.divw in awards:
		awards[hum.divw].append( hum )
	else:
		awards[hum.divw] = [hum]

for key in awards.keys():
	winners = dict()
	lifters = awards[key]
	print "%s Class %s Event" % key
	for lifter in lifters:
		winners[lifter.place] = lifter

	if 1 in winners:
		print "\tFirst Place: %s wilks %r" % ( winners[1].name, winners[1].wilks )
	if 2 in winners:
		print "\tSecond Place: %s wilks %r" % ( winners[2].name, winners[2].wilks )
	if 3 in winners:
		print "\tThird Place: %s wilks %r" % ( winners[3].name, winners[3].wilks )
	print "\n"
	
raw_input()
